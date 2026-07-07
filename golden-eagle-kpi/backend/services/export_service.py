"""金鹰工单KPI管理 - 导出服务"""
import io
import os
from datetime import datetime
from sqlalchemy.orm import Session
from sqlalchemy import text
from backend.models.work_ticket import WorkTicket
from backend.models.personnel import Personnel
from backend.config import AppConfig


class ExportService:
    """导出Excel报表"""

    @staticmethod
    def export_personnel(project_id: int, month: str, db: Session) -> str:
        """导出人力清单Excel，返回文件路径"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "人力清单"

        # 标题行
        headers = ["工号", "姓名", "职务", "系统角色", "项目", "发起工单数", "是否外包", "状态"]
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # 查询数据
        query = db.query(Personnel)
        if project_id:
            query = query.filter(Personnel.project_id == project_id)

        personnel_list = query.all()

        # 批量查询工单数（消除N+1：一次GROUP BY替代逐人查询）
        if month:
            count_rows = db.execute(text(
                "SELECT initiator_id, COUNT(*) as cnt FROM work_tickets "
                "WHERE initiator_id IS NOT NULL "
                "AND strftime('%Y-%m', create_time) = :month "
                "GROUP BY initiator_id"
            ), {"month": month}).fetchall()
        else:
            count_rows = db.execute(text(
                "SELECT initiator_id, COUNT(*) as cnt FROM work_tickets "
                "WHERE initiator_id IS NOT NULL "
                "GROUP BY initiator_id"
            )).fetchall()
        count_map = {row[0]: row[1] for row in count_rows}

        row_idx = 2
        for person in personnel_list:
            count = count_map.get(person.employee_id, 0)

            data = [
                person.employee_id,
                person.name,
                person.role or "",
                _map_role(person),
                str(person.project_id) if person.project_id else "",
                count,
                "是" if person.is_outsourcing else "否",
                person.status or "",
            ]

            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
            row_idx += 1

        # 列宽
        col_widths = [15, 12, 20, 15, 15, 12, 10, 10]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # 保存
        filename = f"人力清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(str(AppConfig.EXPORTS_DIR), filename)
        wb.save(filepath)

        return filepath

    @staticmethod
    def export_kpi_report(project_id: int, month: str, db: Session) -> str:
        """导出KPI报表Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from backend.services.stats_service import StatsService

        wb = openpyxl.Workbook()

        # Sheet1: 综合评分
        ws1 = wb.active
        ws1.title = "综合评分"
        score = StatsService.get_score(project_id, month, db)

        headers = ["指标", "得分", "权重", "加权得分"]
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        items = [
            ("完成率", score["completionScore"], score["weights"]["completion"],
             round(score["completionScore"] * score["weights"]["completion"] / 100, 2)),
            ("及时率", score["timelinessScore"], score["weights"]["timeliness"],
             round(score["timelinessScore"] * score["weights"]["timeliness"] / 100, 2)),
            ("发起达成率", score["initiationScore"], score["weights"]["initiation"],
             round(score["initiationScore"] * score["weights"]["initiation"] / 100, 2)),
        ]

        for row_idx, (name, s, w, ws_val) in enumerate(items, 2):
            for col, val in enumerate([name, s, f"{w}%", ws_val], 1):
                cell = ws1.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

        # 总分
        ws1.cell(row=5, column=1, value="综合评分").font = Font(bold=True)
        ws1.cell(row=5, column=2, value=score["overallScore"]).font = Font(bold=True, size=14)
        ws1.cell(row=6, column=1, value="等级").font = Font(bold=True)
        ws1.cell(row=6, column=2, value=score["grade"]).font = Font(bold=True, size=14, color="FF0000")

        # Sheet2: 四层级统计
        ws2 = wb.create_sheet("四层级统计")
        initiation = StatsService.get_initiation_by_level(project_id, month, db)

        headers2 = ["层级", "人数", "已发起", "目标", "人均目标", "达成率"]
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, level in enumerate(initiation.get("levels", []), 2):
            data = [
                level["level"],
                level.get("count", ""),
                level["initiated"],
                level["target"],
                level.get("targetPerPerson", ""),
                f"{level['achievementRate']}%",
            ]
            for col, val in enumerate(data, 1):
                ws2.cell(row=row_idx, column=col, value=val)

        # Sheet3: 预警清单
        ws3 = wb.create_sheet("预警清单")
        warnings = StatsService.get_warnings(project_id, "", 100, month, db)

        headers3 = ["姓名", "层级", "职务", "已发起", "目标(动态)", "达成率", "预警类型"]
        for col, header in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, item in enumerate(warnings.get("items", []), 2):
            data = [
                item["name"],
                item["level"],
                item.get("position", ""),
                item["initiated"],
                item.get("targetDynamic", ""),
                f"{item['achievementRate']}%",
                "严重" if item["warningType"] == "severe" else "一般",
            ]
            for col, val in enumerate(data, 1):
                ws3.cell(row=row_idx, column=col, value=val)

        filename = f"KPI报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(str(AppConfig.EXPORTS_DIR), filename)
        wb.save(filepath)

        return filepath


    @staticmethod
    def export_tickets(filters: dict, db: Session) -> str:
        """导出工单明细Excel（带KPI状态四色标签）"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from backend.services.ticket_service import TicketService
        from datetime import datetime as dt

        now = dt.now()

        def parse_dt(s):
            if not s:
                return None
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f"):
                try:
                    return dt.strptime(s, fmt)
                except Exception:
                    pass
            return None

        def _calc_kpi_status(create_time_str, complete_time_str, ref_now):
            """计算KPI状态（与前端 computeTicketStatusBadge 逻辑一致）"""
            if not create_time_str:
                return "未知"
            ct = parse_dt(create_time_str)
            if not ct:
                return "未知"
            pt = parse_dt(complete_time_str)
            if pt:
                days = int((pt - ct).total_seconds() / 86400)
                return "已完成" if days <= 7 else "逾期完成"
            else:
                days = int((ref_now - ct).total_seconds() / 86400)
                return "逾期未完成" if days > 7 else "进行中"

        # 复用工单搜索逻辑获取数据
        raw_tickets = TicketService.search_tickets(
            project_id=filters.get("project_id"),
            status=filters.get("status"),
            keyword=filters.get("keyword"),
            start_date=filters.get("start_date"),
            end_date=filters.get("end_date"),
            month=filters.get("month"),
            ticket_type=filters.get("ticket_type"),
            page_size=99999,
            db=db
        )
        tickets = raw_tickets if isinstance(raw_tickets, list) else (raw_tickets.get("items", []) if isinstance(raw_tickets, dict) else [])
        
        # 支持 kpiStatus 筛选（后端实时计算）
        kpi_status_filter = filters.get("kpi_status")
        if kpi_status_filter:
            filtered_tickets = []
            for t in tickets:
                create_time = t.get("createTime")
                complete_time = t.get("completeTime")
                kpi_status = _calc_kpi_status(create_time, complete_time, now)
                if kpi_status in kpi_status_filter:
                    filtered_tickets.append(t)
            tickets = filtered_tickets

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "工单明细"

        # 简化表头：只保留关键字段
        headers = [
            "工单编号", "项目", "类型", "报修描述",
            "发起人", "处理人",
            "创建时间", "完成时间",
            "当前状态", "KPI状态"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        row_idx = 2
        for t in tickets:
            create_time = t.get("createTime")
            complete_time = t.get("completeTime")

            # 计算KPI状态
            kpi_status = _calc_kpi_status(create_time, complete_time, now)

            row = [
                t.get("id") or t.get("ticket_no", ""),
                t.get("projectName", ""),
                t.get("ticketType", ""),
                (t.get("description") or "")[:60],
                t.get("creator", ""),
                t.get("handler", ""),
                (create_time or "")[:16],
                (complete_time or "")[:16],
                t.get("statusText", "") or t.get("orderStatus", ""),
                kpi_status,
            ]
            for col, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col, value=val)

            # 仅 KPI状态 列（第10列）上色
            kpi_fill = {
                "已完成": PatternFill(start_color="E6FFFA", end_color="E6FFFA", fill_type="solid"),
                "逾期完成": PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
                "进行中": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
                "逾期未完成": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
            }.get(kpi_status)
            if kpi_fill:
                ws.cell(row=row_idx, column=10).fill = kpi_fill

            row_idx += 1

        # 列宽（简洁）
        for i, w in enumerate([20, 18, 10, 40, 12, 12, 20, 20, 12, 14], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        filename = f"工单明细_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(str(AppConfig.EXPORTS_DIR), filename)
        wb.save(filepath)
        return filepath


def _map_role(person) -> str:
    """角色映射"""
    if not person.role:
        return "一线员工"
    pos = person.role
    if "总监" in pos or "总经理" in pos:
        return "项目负责人"
    if any(k in pos for k in ["经理", "主管"]):
        return "部门管理"
    if "外包" in pos:
        return "外包"
    return "一线员工"
