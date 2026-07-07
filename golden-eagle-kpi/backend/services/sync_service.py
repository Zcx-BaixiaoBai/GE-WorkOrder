"""金鹰工单KPI管理 - 数据同步服务

流程：启动BI爬虫 → 下载两个Excel → 并发入库（工单+随手拍同时写） → 刷新映射
基于 v0.0.8 稳定版重写：
- fetch_all 只登录一次，串行下载两个报表
- 两个Excel都下载完后，用线程池并发入库
- 随手拍导入使用普通模式读取（处理合并单元格）
"""
import asyncio
import os
import shutil
import traceback
import json
import threading
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor
from sqlalchemy.orm import Session
from sqlalchemy import text
from backend.models.sync_log import SyncLog
from backend.models.work_ticket import WorkTicket
from backend.models.snapshot import Snapshot
from backend.models.project_name_mapping import ProjectNameMapping
from backend.models.project import Project
from backend.config import AppConfig
from backend.database import get_session_local


_sync_status = {
    "is_syncing": False,
    "current_task": None,
    "progress": 0,
    "message": "",
    "last_sync_time": None,
    "last_sync_result": None,
}

# 状态读写锁：多用户并发安全
_sync_lock = threading.Lock()
# 进度更新锁：防止两个入库线程并发写 progress 导致进度条来回跳
_progress_lock = threading.Lock()


class SyncService:

    @staticmethod
    def get_sync_status() -> dict:
        with _sync_lock:
            last_time = _sync_status["last_sync_time"]
            current = dict(_sync_status)  # 浅拷贝避免外部修改
        if not last_time:
            db = get_session_local()()
            try:
                last_time = SyncService.get_last_sync_time(db)
            finally:
                db.close()
        return {
            "isSyncing": current["is_syncing"],
            "currentTask": current["current_task"],
            "progress": _sync_status["progress"],
            "message": _sync_status["message"],
            "lastSyncTime": last_time,
            "lastSyncResult": _sync_status["last_sync_result"],
            "hasError": _sync_status["last_sync_result"] == "failed",
        }

    @staticmethod
    def get_last_sync_time(db: Session) -> str | None:
        row = db.execute(
            text("SELECT finished_at FROM sync_logs WHERE status='completed' ORDER BY id DESC LIMIT 1")
        ).fetchone()
        return row[0] if row and row[0] else None

    @staticmethod
    async def _run_sync_task(account: str, password: str, project_id: int):
        """后台同步任务

        进度分配：
          0-5%   登录BI
          5-45%  下载两个报表（串行）
          45-90% 并发入库
          90-100 刷新映射
        """
        try:
            from backend.scraper.bi_client import BiClient
        except ImportError as e:
            _sync_status["is_syncing"] = False
            _sync_status["progress"] = 0
            _sync_status["message"] = f"同步功能不可用：{e}"
            _sync_status["last_sync_result"] = "failed"
            return

        # Playwright 环境预检
        try:
            from playwright.async_api import async_playwright
            from playwright._impl._driver import compute_driver_executable
            _info = compute_driver_executable()
            _exe = _info[0] if isinstance(_info, tuple) else _info
            if not Path(_exe).exists():
                raise FileNotFoundError(f"Playwright node 不存在: {_exe}")
        except ImportError as e:
            _sync_status.update({"is_syncing": False, "progress": 0, "message": f"Playwright组件缺失: {e}", "last_sync_result": "failed"})
            return
        except FileNotFoundError as e:
            _sync_status.update({"is_syncing": False, "progress": 0, "message": str(e), "last_sync_result": "failed"})
            return
        except Exception as e:
            _sync_status.update({"is_syncing": False, "progress": 0, "message": f"Playwright检查失败: {e}", "last_sync_result": "failed"})
            return

        # 标记同步开始（贯穿整个任务生命周期，finally中重置为False）
        _sync_status["is_syncing"] = True
        _sync_status["progress"] = 0
        _sync_status["last_sync_result"] = None

        db = get_session_local()()
        try:
            sync_log = SyncLog(sync_type="full", project_id=project_id, status="running", started_at=datetime.now())
            db.add(sync_log); db.commit(); db.refresh(sync_log)
            _sync_status["current_task"] = sync_log.id

            # === 阶段1: 登录+下载 (0-45%) ===
            _sync_status["message"] = "正在登录BI系统..."
            _sync_status["progress"] = 2

            client = BiClient(account, password)

            # 总超时保护：整个同步任务最多5分钟
            async def _fetch_with_timeout():
                return await asyncio.wait_for(client.fetch_all(), timeout=300)

            files = await _fetch_with_timeout()

            if not files or len(files) == 0:
                raise RuntimeError("未下载到任何Excel文件，请检查BI系统是否可访问")

            # 检查下载的文件是否有效（非空文件）
            valid_files = []
            for f in files:
                if f and Path(f).exists() and os.path.getsize(f) > 10000:
                    valid_files.append(f)
                else:
                    print(f"[同步] 文件无效或过小，跳过: {f}")

            if not valid_files:
                raise RuntimeError("下载的Excel文件均为空或损坏，同步终止")

            tickets_file = valid_files[0]
            snapshots_file = valid_files[1] if len(valid_files) > 1 else None

            if tickets_file:
                _sync_status["message"] = f"工单明细: {Path(tickets_file).name}"
            if snapshots_file:
                _sync_status["message"] = f"随手拍: {Path(snapshots_file).name}"
            _sync_status["progress"] = 45

            # === 阶段2: 并发入库 (45-90%) ===
            # 工单+随手拍并发执行，各自独立DB session，避免SQLite WAL锁冲突
            _sync_status["message"] = "正在入库数据..."
            _sync_status["progress"] = 48

            ticket_result = {"imported": 0, "skipped": 0}
            snapshot_result = {"imported": 0, "skipped": 0}

            def _import_tickets():
                nonlocal ticket_result
                if not tickets_file:
                    return
                tdb = get_session_local()()
                try:
                    ticket_result = SyncService._import_tickets_from_excel(
                        tickets_file, sync_log.id, tdb,
                        progress_cb=lambda p: _update_progress(48, 70, p)
                    )
                finally:
                    tdb.close()

            def _import_snapshots():
                nonlocal snapshot_result
                if not snapshots_file:
                    return
                sdb = get_session_local()()
                try:
                    snapshot_result = SyncService._import_snapshots_from_excel(
                        snapshots_file, sync_log.id, sdb,
                        progress_cb=lambda p: _update_progress(48, 90, p)
                    )
                finally:
                    sdb.close()

            loop = asyncio.get_event_loop()
            with ThreadPoolExecutor(max_workers=2) as executor:
                futures = []
                if tickets_file:
                    futures.append(loop.run_in_executor(executor, _import_tickets))
                if snapshots_file:
                    futures.append(loop.run_in_executor(executor, _import_snapshots))
                if futures:
                    await asyncio.gather(*futures)

            _sync_status["progress"] = 90

            # === 阶段3: 刷新映射 (90-100%) ===
            _sync_status["message"] = "正在更新项目名称映射..."
            SyncService._refresh_name_mappings(db)
            _sync_status["progress"] = 100
            _sync_status["message"] = "同步完成"
            _sync_status["last_sync_result"] = "completed"

            sync_log.status = "completed"
            sync_log.finished_at = datetime.now()
            sync_log.tickets_synced = ticket_result["imported"]
            sync_log.snapshots_synced = snapshot_result["imported"]
            db.commit()
            _sync_status["last_sync_time"] = datetime.now().isoformat()

        except Exception as e:
            _sync_status["message"] = f"同步失败: {str(e)}"
            _sync_status["last_sync_result"] = "failed"
            traceback.print_exc()
            if _sync_status.get("current_task"):
                log = db.query(SyncLog).filter(SyncLog.id == _sync_status["current_task"]).first()
                if log:
                    log.status = "failed"
                    log.finished_at = datetime.now()
                    log.error_message = str(e)
                    db.commit()
        finally:
            try:
                db.close()
            except Exception:
                pass
            _sync_status["is_syncing"] = False
            _sync_status["current_task"] = None
            try:
                asyncio.get_event_loop().call_later(5, _reset_sync_progress)
            except Exception:
                pass

    @staticmethod
    def _import_tickets_from_excel(file_path: str, batch_id: int, db: Session,
                                    progress_cb=None) -> dict:
        """工单明细导入 - 优化版：单遍流式读取 + 大batch"""
        import openpyxl
        print(f"[同步] 读取工单明细: {Path(file_path).name}")

        mappings = db.query(ProjectNameMapping).all()
        name_map = {m.bi_name: m.standard_name for m in mappings}
        projects = db.query(Project).all()
        projects_by_name = {p.name: p.id for p in projects}

        deleted = db.query(WorkTicket).filter(WorkTicket.source == "detail").delete()
        db.commit()
        print(f"[同步] 已清除旧工单明细 {deleted} 条")
        if progress_cb: progress_cb(0.05)

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        imported = skipped = 0
        batch = []
        report_interval = 1000
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or len(row) < 10:
                    skipped += 1; continue
                ticket_no = str(row[2] or "").strip()
                if not ticket_no:
                    skipped += 1; continue

                raw_project = str(row[1] or "").strip()
                standard_name = name_map.get(raw_project, raw_project)
                project_id = projects_by_name.get(standard_name)
                brand = str(row[4] or "").strip()
                order_type = str(row[7] or "").strip()
                current_node = str(row[9] or "").strip()
                if current_node == "已解决": order_status = "已完成"
                elif current_node in ("待处理", "待派单", "待接单"): order_status = "待处理"
                else: order_status = current_node or "处理中"

                initiator_id = _clean_id(row[11])
                initiator_name = str(row[12] or "").strip()
                create_time = _parse_datetime(row[5])
                complete_time = _parse_datetime(row[20])
                deadline = _parse_datetime(row[6])
                accept_time = _parse_datetime(row[14])
                area_name = str(row[3] or "").strip()
                description = str(row[8] or "").strip()

                ticket_type = None
                if brand == '秩序报修': ticket_type = '秩序报修'
                elif brand == '保洁报修': ticket_type = '保洁报修'

                batch.append(WorkTicket(
                    ticket_no=ticket_no, project_name=raw_project, standard_name=standard_name,
                    project_id=project_id, order_type=order_type, order_status=order_status,
                    initiator_id=initiator_id, initiator_name=initiator_name,
                    create_time=create_time, accept_time=accept_time, complete_time=complete_time,
                    deadline=deadline, area_name=area_name,
                    description=description[:500] if description else None,
                    sync_batch_id=batch_id, source="detail", ticket_type=ticket_type, brand=brand,
                ))
                imported += 1
                if len(batch) >= 1000:
                    db.bulk_save_objects(batch); db.commit(); batch.clear()
                    if progress_cb: progress_cb(0.05 + 0.90 * imported / max(imported + skipped, 1))
                    if imported % report_interval == 0:
                        print(f"  ...工单已导入 {imported} 条")
            except Exception as e:
                skipped += 1
                if skipped <= 10: print(f"  行{row_idx}异常: {e}")
        if batch:
            db.bulk_save_objects(batch); db.commit()
        wb.close()
        if progress_cb: progress_cb(1.0)
        print(f"[同步] 工单导入完成: 导入{imported}, 跳过{skipped}")
        return {"imported": imported, "skipped": skipped}

    @staticmethod
    def _import_snapshots_from_excel(file_path: str, batch_id: int, db: Session,
                                      progress_cb=None) -> dict:
        """随手拍导入 - 优化版：read_only读取数据 + 普通模式读merged_cells

        性能优化：
        - 用普通模式只读 merged_cells（不读数据），构建span map
        - 用 read_only 模式流式读取数据行，内存占用低
        - 合并单元格值从span map中查找
        - batch_size 5000，减少commit次数
        """
        import openpyxl
        print(f"[同步] 读取随手拍: {Path(file_path).name}")

        # Step 1: 普通模式只读 merged_cells 范围，构建 span map
        print("[同步] 扫描合并单元格...")
        wb_meta = openpyxl.load_workbook(file_path, data_only=True)
        ws_meta = wb_meta.active
        merged_spans = {}  # (row, col) -> top_left_value
        for merged_range in ws_meta.merged_cells.ranges:
            top_row, bottom_row = merged_range.min_row, merged_range.max_row
            left_col, right_col = merged_range.min_col, merged_range.max_col
            top_left_val = ws_meta.cell(row=top_row, column=left_col).value
            for r in range(top_row, bottom_row + 1):
                for c in range(left_col, right_col + 1):
                    if (r, c) != (top_row, left_col):
                        merged_spans[(r, c)] = top_left_val
        wb_meta.close()

        mappings = db.query(ProjectNameMapping).all()
        name_map = {m.bi_name: m.standard_name for m in mappings}
        projects = db.query(Project).all()
        projects_by_name = {p.name: p.id for p in projects}
        for p in projects:
            if p.bi_names:
                bi_list = json.loads(p.bi_names) if isinstance(p.bi_names, str) else p.bi_names
                for bn in bi_list:
                    if bn not in projects_by_name:
                        projects_by_name[bn] = p.id

        # 暴力覆盖
        snap_deleted = db.query(Snapshot).delete()
        db.commit()
        print(f"[同步] 已清除旧随手拍: snapshots={snap_deleted}")
        if progress_cb: progress_cb(0.05)

        # Step 2: read_only 流式读取数据
        print("[同步] 开始流式导入随手拍...")
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        imported = skipped = 0
        snap_batch = []
        row_count = 0
        report_interval = 1000  # 每1000条报告一次（更频繁更新进度）

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_count += 1
            try:
                def _cell(col):
                    """获取单元格值，如果是合并单元格成员则返回左上角值"""
                    val = row[col - 1] if col - 1 < len(row) else None
                    if val is None and (row_idx, col) in merged_spans:
                        val = merged_spans[(row_idx, col)]
                    return val

                if _cell(1) is None and _cell(2) is None:
                    skipped += 1; continue

                ticket_no = f"S{row_idx:08d}"
                initiator_id = _clean_id(_cell(1))
                initiator_name = str(_cell(2) or "").strip()
                create_time = _parse_datetime(_cell(6))
                problem_type = str(_cell(7) or "").strip()

                raw_status = str(_cell(12) or "").strip()
                if raw_status == "已解决": order_status = "已完成"
                elif raw_status in ("待处理", "未处理"): order_status = "待处理"
                else: order_status = raw_status or "处理中"

                complete_time = _parse_datetime(_cell(13))

                raw_project = standard_name = ""
                project_id = None
                dept = str(_cell(3) or "").strip()
                if dept and '/' in dept:
                    possible_name = dept.split('/')[0].strip()
                    mapped = name_map.get(possible_name)
                    if mapped:
                        raw_project, standard_name = possible_name, mapped
                        project_id = projects_by_name.get(standard_name)
                    else:
                        for p in projects:
                            if p.name in possible_name or possible_name in p.name:
                                standard_name, project_id, raw_project = p.name, p.id, possible_name
                                break

                description = str(_cell(8) or "").strip()

                snap_batch.append(Snapshot(
                    ticket_no=ticket_no, project_name=raw_project, standard_name=standard_name,
                    project_id=project_id, order_type=problem_type, order_status=order_status,
                    initiator_id=initiator_id, initiator_name=initiator_name,
                    create_time=create_time, complete_time=complete_time,
                    area_name=dept[:50] if dept else None,
                    description=description[:500] if description else None,
                    sync_batch_id=batch_id,
                ))
                imported += 1
                if len(snap_batch) >= 1000:
                    db.bulk_save_objects(snap_batch); db.commit(); snap_batch.clear()
                    if progress_cb:
                        progress_cb(0.05 + 0.9 * imported / max(imported + skipped, max(row_count, 1)))
                    if imported % report_interval == 0:
                        print(f"  ...随手拍已导入 {imported} 条")
            except Exception as e:
                skipped += 1
                if skipped <= 10: print(f"  行{row_idx}异常: {e}")

        if snap_batch:
            db.bulk_save_objects(snap_batch); db.commit()
        wb.close()
        if progress_cb: progress_cb(1.0)
        print(f"[同步] 随手拍导入完成: 导入{imported}, 跳过{skipped}")
        return {"imported": imported, "skipped": skipped}

    @staticmethod
    def _refresh_name_mappings(db: Session):
        raw_names = db.execute(text(
            "SELECT DISTINCT project_name FROM work_tickets WHERE project_name IS NOT NULL"
        )).fetchall()
        for (raw_name,) in raw_names:
            existing = db.query(ProjectNameMapping).filter(ProjectNameMapping.bi_name == raw_name).first()
            if not existing:
                db.add(ProjectNameMapping(bi_name=raw_name, standard_name=raw_name, source="bi"))
        db.commit()

    @staticmethod
    def get_sync_logs(page: int = 1, page_size: int = 10, db: Session = None) -> dict:
        query = db.query(SyncLog).order_by(SyncLog.started_at.desc())
        total = query.count()
        items = query.offset((page - 1) * page_size).limit(page_size).all()
        return {
            "total": total, "page": page, "pageSize": page_size,
            "items": [
                {"id": log.id, "syncType": log.sync_type,
                 "projectId": str(log.project_id) if log.project_id else None,
                 "status": log.status,
                 "startedAt": log.started_at.isoformat() if log.started_at else None,
                 "completedAt": log.finished_at.isoformat() if log.finished_at else None,
                 "ticketsSynced": log.tickets_synced, "snapshotsSynced": log.snapshots_synced,
                 "errorMessage": log.error_message}
                for log in items
            ],
        }


def _clean_id(raw_id) -> str:
    if raw_id is None: return "0000000000"
    if isinstance(raw_id, float): raw_id = str(int(raw_id))
    raw_id = str(raw_id).strip()
    if not raw_id or raw_id == 'None': return "0000000000"
    return raw_id.zfill(10)


def _update_progress(start_pct: int, end_pct: int, internal_pct: float):
    global _sync_status
    with _progress_lock:
        p = start_pct + (end_pct - start_pct) * internal_pct
        new_val = int(min(p, end_pct))
        # 进度只进不退，防止两个入库线程互相覆盖导致进度条来回跳
        if new_val > _sync_status["progress"]:
            _sync_status["progress"] = new_val


def _parse_datetime(val):
    if val is None: return None
    if isinstance(val, datetime): return val
    s = str(val).strip()
    if not s or s in ('None', 'NaT', ''): return None
    if len(s) == 14 and s.isdigit(): return datetime.strptime(s, "%Y%m%d%H%M%S")
    if len(s) == 8 and s.isdigit(): return datetime.strptime(s, "%Y%m%d")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
        try: return datetime.strptime(s, fmt)
        except ValueError: continue
    return None


def _reset_sync_progress():
    _sync_status["progress"] = 0
