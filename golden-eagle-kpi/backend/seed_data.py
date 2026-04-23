"""金鹰工单KPI管理 - 初始化数据导入脚本

从三个Excel文件读取数据灌入SQLite数据库：
1. BI项目名称对照.xlsx → projects + project_name_mapping
2. 金鹰物业项目负责人清单.xlsx → projects.manager_name + role_mapping
3. 初始化人力清单.xlsx → project_name_mapping（人力库名称对照）

运行方式: python -m backend.seed_data
"""
import json
import sys
from pathlib import Path

import openpyxl

# 数据文件路径
MEMORY_DIR = Path(r"C:\Users\Administrator\记忆\初始化数据集")
BI_MAPPING_FILE = MEMORY_DIR / "BI项目名称对照.xlsx"
MANAGER_FILE = MEMORY_DIR / "金鹰物业项目负责人清单.xlsx"
PERSONNEL_FILE = MEMORY_DIR / "初始化人力清单.xlsx"


def read_bi_mapping():
    """读取BI项目名称对照表"""
    wb = openpyxl.load_workbook(str(BI_MAPPING_FILE), data_only=True)
    ws = wb.active
    # Headers: [序号, 标准项目名称, 项目属性, 去除空置考核面积（万方）, BI—项目名称(col5), col6, col7]
    projects = []
    mappings = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        seq = row[0]
        standard_name = row[1]
        project_type = row[2]  # 项目属性：商写酒医/住 等
        area_wanfang = row[3]  # 万方

        if not standard_name:
            continue

        # 面积：万方 → 平方米
        area_m2 = None
        if area_wanfang is not None:
            try:
                area_m2 = float(area_wanfang) * 10000
            except (ValueError, TypeError):
                pass

        # BI名称列表（第5、6、7列，可能为空）
        bi_names = []
        for col_idx in [4, 5, 6]:
            val = row[col_idx] if col_idx < len(row) else None
            if val and str(val).strip():
                bi_names.append(str(val).strip())

        projects.append({
            "seq": seq,
            "name": standard_name,
            "project_type": project_type,
            "area": area_m2,
            "bi_names": json.dumps(bi_names, ensure_ascii=False) if bi_names else None,
        })

        # 映射关系
        for bi_name in bi_names:
            mappings.append({
                "bi_name": bi_name,
                "standard_name": standard_name,
            })

    wb.close()
    return projects, mappings


def read_personnel_mapping():
    """读取人力清单中的名称对照（第二个sheet）"""
    wb = openpyxl.load_workbook(str(PERSONNEL_FILE), data_only=True)
    # Sheet 2: 名称对照
    if "名称对照" in wb.sheetnames:
        ws = wb["名称对照"]
    else:
        ws = wb.active

    # Headers: [序号, 项目名称, 项目属性, 人力库—项目名称]
    mappings = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        standard_name = row[1]
        hr_name = row[3]  # 人力库名称
        if standard_name and hr_name and str(hr_name).strip():
            hr_name_str = str(hr_name).strip()
            # 跳过与标准名完全相同的（不需要映射）
            if hr_name_str != str(standard_name).strip():
                mappings.append({
                    "standard_name": str(standard_name).strip(),
                    "hr_name": hr_name_str,
                })

    wb.close()
    return mappings


def read_manager_list():
    """读取项目负责人清单"""
    wb = openpyxl.load_workbook(str(MANAGER_FILE), data_only=True)
    ws = wb["负责人清单"]

    managers = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        project_name = row[0]
        manager_name = row[1]
        if project_name and manager_name:
            managers[str(project_name).strip()] = str(manager_name).strip()

    wb.close()
    return managers


def read_role_mapping():
    """读取管理层岗位清单"""
    wb = openpyxl.load_workbook(str(MANAGER_FILE), data_only=True)
    ws = wb["管理层岗位清单"]

    roles = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        role = row[0]
        if role and str(role).strip():
            roles.append(str(role).strip())

    wb.close()
    return roles


def classify_role(role_name: str) -> str:
    """根据角色名推断角色分类"""
    if "总监" in role_name:
        return "管理层-总监级"
    elif "总经理" in role_name:
        return "管理层-总经理级"
    elif "副总" in role_name:
        return "管理层-副总级"
    elif "经理" in role_name and "副" not in role_name:
        return "管理层-经理级"
    elif "副经理" in role_name:
        return "管理层-副经理级"
    elif "主管" in role_name:
        return "管理层-主管级"
    elif "管培生" in role_name:
        return "管培生"
    else:
        return "管理层-其他"


def seed_database():
    """主函数：导入所有初始化数据"""
    from backend.database import init_engine, SessionLocal
    from backend.models.project import Project
    from backend.models.project_name_mapping import ProjectNameMapping
    from backend.models.role_mapping import RoleMapping

    # 初始化引擎
    init_engine()

    # 读取数据
    print("[数据导入] 读取BI项目名称对照表...")
    projects_data, bi_mappings = read_bi_mapping()
    print(f"  → {len(projects_data)} 个项目, {len(bi_mappings)} 条BI映射")

    print("[数据导入] 读取人力库名称对照表...")
    hr_mappings = read_personnel_mapping()
    print(f"  → {len(hr_mappings)} 条人力库映射")

    print("[数据导入] 读取项目负责人清单...")
    managers = read_manager_list()
    print(f"  → {len(managers)} 位项目负责人")

    print("[数据导入] 读取管理层岗位清单...")
    roles = read_role_mapping()
    print(f"  → {len(roles)} 个管理层岗位")

    # 写入数据库
    db = SessionLocal()
    try:
        # 1. 清空旧数据
        print("[数据导入] 清空旧数据...")
        db.query(ProjectNameMapping).delete()
        db.query(RoleMapping).delete()
        db.query(Project).delete()
        db.commit()

        # 2. 导入项目
        print("[数据导入] 导入项目数据...")
        project_map = {}  # name → Project object
        for p in projects_data:
            project = Project(
                name=p["name"],
                bi_names=p["bi_names"],
                area=p["area"],
                manager_name=managers.get(p["name"]),
                kpi_completion_rate=95.0,
                kpi_timely_rate=90.0,
                status="active",
            )
            db.add(project)
            project_map[p["name"]] = project

        db.commit()

        # 刷新获取ID
        for name, project in project_map.items():
            db.refresh(project)

        print(f"  → 已导入 {len(project_map)} 个项目")

        # 3. 导入BI名称映射
        print("[数据导入] 导入BI名称映射...")
        bi_mapping_count = 0
        seen_bi_names = set()
        for m in bi_mappings:
            bi_name = m["bi_name"]
            if bi_name in seen_bi_names:
                continue
            seen_bi_names.add(bi_name)

            project = project_map.get(m["standard_name"])
            mapping = ProjectNameMapping(
                bi_name=bi_name,
                standard_name=m["standard_name"],
                project_id=project.id if project else None,
                source="bi",
            )
            db.add(mapping)
            bi_mapping_count += 1

        # 4. 导入人力库名称映射
        print("[数据导入] 导入人力库名称映射...")
        hr_mapping_count = 0
        for m in hr_mappings:
            hr_name = m["hr_name"]
            if hr_name in seen_bi_names:
                # 已经作为BI映射存在，跳过
                continue
            seen_bi_names.add(hr_name)

            project = project_map.get(m["standard_name"])
            mapping = ProjectNameMapping(
                bi_name=hr_name,
                standard_name=m["standard_name"],
                project_id=project.id if project else None,
                source="hr",
            )
            db.add(mapping)
            hr_mapping_count += 1

        db.commit()
        print(f"  → BI映射 {bi_mapping_count} 条, 人力库映射 {hr_mapping_count} 条")

        # 5. 导入角色映射
        print("[数据导入] 导入管理层岗位映射...")
        for role in roles:
            rm = RoleMapping(
                source_role=role,
                target_role=role,  # 初始1:1映射，后续可调整
                category=classify_role(role),
            )
            db.add(rm)
        db.commit()
        print(f"  → 已导入 {len(roles)} 个角色映射")

        # 6. 验证
        print("\n[数据导入] 验证结果:")
        project_count = db.query(Project).count()
        mapping_count = db.query(ProjectNameMapping).count()
        role_count = db.query(RoleMapping).count()
        projects_with_manager = db.query(Project).filter(Project.manager_name.isnot(None)).count()
        projects_with_area = db.query(Project).filter(Project.area.isnot(None)).count()

        print(f"  项目数: {project_count}")
        print(f"  名称映射数: {mapping_count}")
        print(f"  角色映射数: {role_count}")
        print(f"  有负责人的项目: {projects_with_manager}")
        print(f"  有面积的项目: {projects_with_area}")

        # 打印项目概览
        print("\n[数据导入] 项目概览:")
        all_projects = db.query(Project).order_by(Project.id).all()
        for p in all_projects:
            area_str = f"{p.area/10000:.2f}万㎡" if p.area else "无面积"
            mgr_str = p.manager_name or "无负责人"
            print(f"  #{p.id:2d} {p.name:<20s} {area_str:<12s} 负责人: {mgr_str}")

        print("\n[数据导入] [OK] 数据导入完成!")

    except Exception as e:
        db.rollback()
        print(f"\n[数据导入] [FAIL] 导入失败: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    # 确保可以找到项目模块
    project_root = Path(__file__).resolve().parent.parent
    sys.path.insert(0, str(project_root))

    from backend.database import init_database
    init_database()
    seed_database()
